from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

try:
    from .merge_feedback import FEEDBACK_COLUMNS
    from .features import (
        CATEGORIES,
        RISK_LABELS,
        detect_damage,
        map_risk_label,
        normalize_category,
        parse_is_shared,
        parse_remaining_pct,
        parse_shelf_life,
        parse_used_days,
        parse_user_count,
        parse_weekly_use_count,
        text_or_empty,
    )
except ImportError:  # pragma: no cover - supports direct script execution
    from merge_feedback import FEEDBACK_COLUMNS
    from features import (
        CATEGORIES,
        RISK_LABELS,
        detect_damage,
        map_risk_label,
        normalize_category,
        parse_is_shared,
        parse_remaining_pct,
        parse_shelf_life,
        parse_used_days,
        parse_user_count,
        parse_weekly_use_count,
        text_or_empty,
    )


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = PROJECT_ROOT / "data" / "raw"
PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"
FEEDBACK_DIR = PROJECT_ROOT / "data" / "feedback"
MANUAL_DIR = PROJECT_ROOT / "data" / "manual"
DEFAULT_INPUT = RAW_DIR / "interview_form_desensitized.xlsx"
EXPANDED_BLUEPRINTS_PATH = MANUAL_DIR / "expanded_blueprints.csv"
HOLDOUT_TERMS_PATH = MANUAL_DIR / "holdout_exclusion_terms.csv"

FIELDNAMES = [
    "sample_id",
    "item_name",
    "user_description",
    "category",
    "risk_label",
    "used_days",
    "remaining_pct",
    "weekly_use_count",
    "user_count",
    "is_shared",
    "has_shelf_life",
    "days_to_expire",
    "is_damaged",
    "source",
    "label_source",
    "original_category",
    "original_user_judgment",
    "remaining_text",
    "frequency_text",
    "shelf_life_text",
    "notes",
]

DEFAULT_HOLDOUT_KEYWORDS = [
    "抽纸/纸巾",
    "垃圾袋",
    "洗发水",
    "创可贴",
    "感冒药",
    "蛋白粉",
    "笔芯",
    "充电线",
    "备用药盒",
]

DEFAULT_HOLDOUT_SIMILAR_KEYWORDS = [
    "抽纸",
    "纸巾",
    "面巾纸",
    "垃圾袋",
    "洗发水",
    "创可贴",
    "感冒药",
    "备用药",
    "蛋白粉",
    "补剂",
    "笔芯",
    "中性笔",
    "充电线",
    "数据线",
    "Type-C",
    "type-c",
    "备用药盒",
    "药收纳盒",
    "常用药收纳盒",
]


def main() -> None:
    parser = argparse.ArgumentParser(description="Prepare desensitized data and ML datasets.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Path to the interview workbook.")
    args = parser.parse_args()

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    FEEDBACK_DIR.mkdir(parents=True, exist_ok=True)
    MANUAL_DIR.mkdir(parents=True, exist_ok=True)

    input_path = args.input
    desensitized_path = RAW_DIR / "interview_form_desensitized.xlsx"
    if not input_path.exists():
        raise FileNotFoundError(f"访谈表不存在：{input_path}")
    if input_path.resolve() != desensitized_path.resolve():
        create_desensitized_workbook(input_path, desensitized_path)
        input_path = desensitized_path

    real_samples = read_real_samples(input_path)
    raw_expanded_samples = build_expanded_samples()
    expanded_samples = filter_expanded_leakage(raw_expanded_samples)
    all_samples = real_samples + expanded_samples

    train_samples, test_samples = split_train_test(real_samples, expanded_samples)
    assert_no_rule_initial(train_samples)

    write_csv(PROCESSED_DIR / "real_samples.csv", real_samples)
    write_csv(PROCESSED_DIR / "expanded_samples.csv", expanded_samples)
    write_csv(PROCESSED_DIR / "all_samples.csv", all_samples)
    write_csv(PROCESSED_DIR / "train.csv", train_samples)
    write_csv(PROCESSED_DIR / "test_real_holdout.csv", test_samples)
    ensure_feedback_log()

    print(f"real_samples={len(real_samples)}")
    print(f"expanded_samples={len(expanded_samples)}")
    print(f"filtered_expanded_samples={len(raw_expanded_samples) - len(expanded_samples)}")
    print(f"train_samples={len(train_samples)}")
    print(f"test_real_holdout={len(test_samples)}")


def create_desensitized_workbook(input_path: Path, output_path: Path) -> None:
    source = load_workbook(input_path, data_only=True)
    target = Workbook()
    target.remove(target.active)

    for source_ws in source.worksheets:
        ws = target.create_sheet(source_ws.title)
        for row in source_ws.iter_rows(values_only=True):
            values = list(row)
            if source_ws.title == "用户基本信息" and values:
                field = text_or_empty(values[0])
                if field == "姓名或称呼" and len(values) > 1:
                    values[1] = "目标用户A"
                if field == "联系方式" and len(values) > 1:
                    values[1] = "[已脱敏]"
            ws.append(values)

    target.save(output_path)


def read_real_samples(input_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    ws = workbook["真实物品样本"]
    samples: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=6, values_only=True):
        sequence = text_or_empty(row[0])
        item_name = text_or_empty(row[1])
        if not item_name or sequence == "示例":
            continue
        description = text_or_empty(row[2])
        original_category = text_or_empty(row[3])
        usage_text = text_or_empty(row[4])
        remaining_text = text_or_empty(row[5])
        frequency_text = text_or_empty(row[6])
        user_count_text = text_or_empty(row[7])
        shelf_life_text = text_or_empty(row[8])
        user_judgment = text_or_empty(row[9])
        notes = text_or_empty(row[12])

        category = normalize_category(original_category, item_name, description)
        risk_label = map_risk_label(user_judgment)
        used_days = parse_used_days(usage_text)
        remaining_pct = parse_remaining_pct(remaining_text)
        weekly_use_count = parse_weekly_use_count(frequency_text)
        user_count = parse_user_count(user_count_text)
        is_shared = parse_is_shared(user_count_text)
        has_shelf_life, days_to_expire = parse_shelf_life(shelf_life_text)
        is_damaged = detect_damage(item_name, description, remaining_text, shelf_life_text, notes)

        samples.append(
            make_sample(
                sample_id=f"real_{int(float(sequence)):03d}",
                item_name=item_name,
                user_description=description,
                category=category,
                risk_label=risk_label,
                used_days=used_days,
                remaining_pct=remaining_pct,
                weekly_use_count=weekly_use_count,
                user_count=user_count,
                is_shared=is_shared,
                has_shelf_life=has_shelf_life,
                days_to_expire=days_to_expire,
                is_damaged=is_damaged,
                source="real_user",
                label_source="user_interview",
                original_category=original_category,
                original_user_judgment=user_judgment,
                remaining_text=remaining_text,
                frequency_text=frequency_text,
                shelf_life_text=shelf_life_text,
                notes=notes,
            )
        )

    return samples


def build_expanded_samples() -> list[dict[str, Any]]:
    if not EXPANDED_BLUEPRINTS_PATH.exists():
        raise FileNotFoundError(
            "未找到人工扩展样本表：data/manual/expanded_blueprints.csv。"
            "请保留该文件，或先从当前项目数据重新生成。"
        )

    samples: list[dict[str, Any]] = []
    with EXPANDED_BLUEPRINTS_PATH.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for index, row in enumerate(reader, start=1):
            sample_id = text_or_empty(row.get("sample_id")) or f"expanded_{index:03d}"
            category = normalize_category(row.get("category", ""), row.get("item_name", ""), row.get("user_description", ""))
            risk_label = text_or_empty(row.get("risk_label"))
            samples.append(
                make_sample(
                    sample_id=sample_id,
                    item_name=text_or_empty(row.get("item_name")),
                    user_description=text_or_empty(row.get("user_description")),
                    category=category,
                    risk_label=risk_label,
                    used_days=to_int(row.get("used_days"), 30),
                    remaining_pct=to_float(row.get("remaining_pct"), 50.0),
                    weekly_use_count=to_float(row.get("weekly_use_count"), 1.0),
                    user_count=to_int(row.get("user_count"), 1),
                    is_shared=to_int(row.get("is_shared"), 0),
                    has_shelf_life=to_int(row.get("has_shelf_life"), 0),
                    days_to_expire=to_int(row.get("days_to_expire"), 999),
                    is_damaged=to_int(row.get("is_damaged"), 0),
                    source=text_or_empty(row.get("source")) or "expanded_manual",
                    label_source=text_or_empty(row.get("label_source")) or "manual_review",
                    original_category=text_or_empty(row.get("original_category")) or category,
                    original_user_judgment=text_or_empty(row.get("original_user_judgment")) or risk_label,
                    remaining_text=text_or_empty(row.get("remaining_text")) or f"{to_float(row.get('remaining_pct'), 50.0)}%",
                    frequency_text=text_or_empty(row.get("frequency_text")) or f"每周约{to_float(row.get('weekly_use_count'), 1.0)}次",
                    shelf_life_text=text_or_empty(row.get("shelf_life_text")) or ("有保质期" if to_int(row.get("has_shelf_life"), 0) else "无"),
                    notes=text_or_empty(row.get("notes")) or "人工扩展并复核的宿舍常见物品样本。",
                )
            )
    return samples


def make_sample(**kwargs: Any) -> dict[str, Any]:
    sample = {field: kwargs.get(field, "") for field in FIELDNAMES}
    if sample["category"] not in CATEGORIES:
        raise ValueError(f"Unknown category: {sample['category']}")
    if sample["risk_label"] not in RISK_LABELS:
        raise ValueError(f"Unknown risk label: {sample['risk_label']}")
    return sample


def split_train_test(
    real_samples: list[dict[str, Any]], expanded_samples: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    test_samples: list[dict[str, Any]] = []
    real_train: list[dict[str, Any]] = []
    for sample in real_samples:
        if is_holdout(sample["item_name"]):
            test_samples.append(sample)
        else:
            real_train.append(sample)
    return real_train + expanded_samples, test_samples


def load_holdout_terms() -> tuple[list[str], list[str]]:
    if not HOLDOUT_TERMS_PATH.exists():
        return DEFAULT_HOLDOUT_KEYWORDS, DEFAULT_HOLDOUT_SIMILAR_KEYWORDS

    holdout_terms: list[str] = []
    similar_terms: list[str] = []
    with HOLDOUT_TERMS_PATH.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row in reader:
            term = text_or_empty(row.get("term"))
            term_type = text_or_empty(row.get("type"))
            if not term:
                continue
            if term_type == "holdout":
                holdout_terms.append(term)
            elif term_type == "similar":
                similar_terms.append(term)

    return holdout_terms or DEFAULT_HOLDOUT_KEYWORDS, similar_terms or DEFAULT_HOLDOUT_SIMILAR_KEYWORDS


def is_holdout(item_name: str) -> bool:
    holdout_terms, _ = load_holdout_terms()
    return any(keyword in item_name for keyword in holdout_terms)


def filter_expanded_leakage(samples: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [sample for sample in samples if not is_holdout_similar(sample["item_name"])]


def is_holdout_similar(item_name: str) -> bool:
    _, similar_terms = load_holdout_terms()
    compact = item_name.replace(" ", "")
    return any(keyword in compact for keyword in similar_terms)


def assert_no_rule_initial(samples: list[dict[str, Any]]) -> None:
    offenders = [sample["sample_id"] for sample in samples if sample["label_source"] == "rule_initial"]
    if offenders:
        joined = ", ".join(offenders[:10])
        raise ValueError(f"Samples with label_source=rule_initial cannot enter training: {joined}")


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def ensure_feedback_log() -> None:
    path = FEEDBACK_DIR / "user_feedback_log.csv"
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        with path.open("r", newline="", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file)
            rows = [{field: row.get(field, "") for field in FEEDBACK_COLUMNS} for row in reader]
            if reader.fieldnames == FEEDBACK_COLUMNS:
                return
        with path.open("w", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=FEEDBACK_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
        return
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=FEEDBACK_COLUMNS)
        writer.writeheader()


def to_float(value: Any, default: float) -> float:
    text = text_or_empty(value)
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def to_int(value: Any, default: int) -> int:
    return int(to_float(value, float(default)))


if __name__ == "__main__":
    main()
